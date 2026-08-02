#!/usr/bin/env python3
"""
seed_warehouse_data.py — Seeds PostgreSQL with real data from warehouse_database.xlsx.

This script reads the warehouse database spreadsheet and populates:
  1. Demo Warehouse (idempotently if missing)
  2. Warehouse topology (zones, aisles, racks, shelves, bins with qr_code)
  3. Products (from spreadsheet: SKU = Product_Code, synthesized name)
  4. Inventory records (maps each product to its bin with expected_qty = 1)

Usage:
    python scripts/seed_warehouse_data.py

Environment Variables:
    DATABASE_URL  — PostgreSQL connection string
                    Default: postgresql://warehouse_admin:warehouse_secret@localhost:5432/warehouse_platform
    XLSX_PATH     — Optional path to warehouse_database.xlsx
"""
from __future__ import annotations

import os
import sys
import uuid
import re

# Try to use openpyxl for xlsx parsing
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

try:
    import psycopg2
except ImportError:
    print("ERROR: psycopg2 is required. Install it with: pip install psycopg2-binary")
    sys.exit(1)

# ── Configuration ──────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://warehouse_admin:warehouse_secret@localhost:5432/warehouse_platform",
)

# Convert async URL to sync if needed
DATABASE_URL = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")

XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(os.path.dirname(__file__)), "warehouse_database.xlsx")
)

WAREHOUSE_ID = "a1b2c3d4-e5f6-7890-abcd-ef1234567890"  # Must match init.sql demo warehouse


# ── Deterministic UUID generation ──────────────────────────────────────────────
def make_uuid(namespace: str, name: str) -> str:
    """Generate a deterministic UUID v5 from a namespace and name."""
    return str(uuid.uuid5(uuid.UUID("6ba7b810-9dad-11d1-80b4-00c04fd430c8"), f"{namespace}:{name}"))


# ── Helper for header lookup ───────────────────────────────────────────────────
def get_field(record: dict, *keys: str, default: str = "") -> str:
    """Extract value from record using candidate keys (handles exact and normalized names)."""
    for key in keys:
        if key in record and record[key] is not None:
            val = str(record[key]).strip()
            if val:
                return val
        normalized_target = key.lower().replace(" ", "").replace("_", "")
        for r_key, r_val in record.items():
            if r_key and r_val is not None:
                normalized_r = str(r_key).lower().replace(" ", "").replace("_", "")
                if normalized_r == normalized_target:
                    val = str(r_val).strip()
                    if val:
                        return val
    return default


# ── Parse XLSX ──────────────────────────────────────────────────────────────────
def parse_warehouse_data(xlsx_path: str) -> list[dict]:
    """Parse the warehouse_database.xlsx using openpyxl and return a list of product records."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = []
    products = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue

        if not any(v is not None for v in row):
            continue

        record = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                record[headers[col_idx]] = value

        products.append(record)

    wb.close()
    return products


def extract_location_from_code(product_code: str) -> dict:
    """
    Extract aisle, row, rack, shelf, product position from product code.
    Expected format: WH-A{aisle}-R{row}-RK{rack}-S{shelf}-P{product}
    """
    pattern = r"WH-A(\d+)-R(\d+)-RK(\d+)-S(\d+)-P(\d+)"
    match = re.match(pattern, product_code)
    if not match:
        return {}
    return {
        "aisle": int(match.group(1)),
        "row": int(match.group(2)),
        "rack": int(match.group(3)),
        "shelf": int(match.group(4)),
        "product_pos": int(match.group(5)),
    }


# ── Seed Functions ──────────────────────────────────────────────────────────────
def seed_warehouse(conn):
    """Ensure the demo warehouse exists in PostgreSQL database idempotently."""
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO warehouses (id, code, name, address, city, country, total_area_sqm, timezone)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (
        WAREHOUSE_ID,
        "WH-001",
        "Primary Distribution Center",
        "123 Industrial Blvd",
        "Bangalore",
        "India",
        25000.00,
        "Asia/Kolkata"
    ))
    conn.commit()
    cur.close()
    print(f"  ✓ Warehouse verified/seeded idempotently: {WAREHOUSE_ID}")


def seed_topology(conn, products: list[dict]) -> dict[str, str]:
    """Create topology entries (zones, aisles, racks, shelves, bins) and store bins.qr_code."""
    cur = conn.cursor()

    aisles_seen = set()
    racks_seen = set()
    shelves_seen = set()
    bins_to_create = []

    for p in products:
        code = get_field(p, "Product_Code", "Product Code", "product_code")
        qr_code = get_field(p, "QR_Code", "QR Code", "qr_code")
        loc = extract_location_from_code(code)
        if not loc:
            continue

        aisles_seen.add(loc["aisle"])
        racks_seen.add((loc["aisle"], loc["rack"]))
        shelves_seen.add((loc["aisle"], loc["rack"], loc["shelf"]))
        bins_to_create.append((loc, qr_code))

    # Create a single zone for the warehouse
    zone_id = make_uuid("zone", f"WH-{WAREHOUSE_ID}-main")
    cur.execute("""
        INSERT INTO zones (id, warehouse_id, code, name, zone_type)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (warehouse_id, code) DO NOTHING
    """, (zone_id, WAREHOUSE_ID, "ZONE-MAIN", "Main Storage Zone", "STORAGE"))

    # Create aisles
    aisle_ids = {}
    for aisle_num in sorted(aisles_seen):
        aisle_id = make_uuid("aisle", f"WH-{WAREHOUSE_ID}-A{aisle_num}")
        aisle_code = f"A{aisle_num}"
        cur.execute("""
            INSERT INTO aisles (id, zone_id, code, aisle_number)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (zone_id, code) DO NOTHING
        """, (aisle_id, zone_id, aisle_code, aisle_num))
        aisle_ids[aisle_num] = aisle_id

    # Create racks
    rack_ids = {}
    for aisle_num, rack_num in sorted(racks_seen):
        rack_id = make_uuid("rack", f"WH-{WAREHOUSE_ID}-A{aisle_num}-RK{rack_num}")
        rack_code = f"A{aisle_num}-RK{rack_num}"
        aisle_id = aisle_ids[aisle_num]
        cur.execute("""
            INSERT INTO racks (id, aisle_id, code, rack_number, num_shelves)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (aisle_id, code) DO NOTHING
        """, (rack_id, aisle_id, rack_code, rack_num, 4))
        rack_ids[(aisle_num, rack_num)] = rack_id

    # Create shelves
    shelf_ids = {}
    for aisle_num, rack_num, shelf_num in sorted(shelves_seen):
        shelf_id = make_uuid("shelf", f"WH-{WAREHOUSE_ID}-A{aisle_num}-RK{rack_num}-S{shelf_num}")
        shelf_code = f"A{aisle_num}-RK{rack_num}-S{shelf_num}"
        rack_id = rack_ids[(aisle_num, rack_num)]
        cur.execute("""
            INSERT INTO shelves (id, rack_id, code, level_number)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (rack_id, level_number) DO NOTHING
        """, (shelf_id, rack_id, shelf_code, shelf_num))
        shelf_ids[(aisle_num, rack_num, shelf_num)] = shelf_id

    # Create bins (one per product position) and set bins.qr_code
    bin_ids = {}
    for loc, qr_code_val in bins_to_create:
        a, rk, s, pos = loc["aisle"], loc["rack"], loc["shelf"], loc["product_pos"]
        bin_code = f"A{a}-RK{rk}-S{s}-B{pos}"
        bin_id = make_uuid("bin", f"WH-{WAREHOUSE_ID}-{bin_code}")
        shelf_id = shelf_ids.get((a, rk, s))
        if not shelf_id:
            continue

        cur.execute("""
            INSERT INTO bins (id, shelf_id, code, bin_number, qr_code)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (code) DO UPDATE SET
                qr_code = EXCLUDED.qr_code,
                updated_at = NOW()
        """, (bin_id, shelf_id, bin_code, pos, qr_code_val if qr_code_val else None))
        bin_ids[bin_code] = bin_id

    conn.commit()
    cur.close()
    print(f"  ✓ Topology seeded: {len(aisles_seen)} aisles, {len(rack_ids)} racks, "
          f"{len(shelf_ids)} shelves, {len(bin_ids)} bins (with qr_code)")
    return bin_ids


def seed_products(conn, products: list[dict]) -> int:
    """Insert products from spreadsheet into products table using Product_Code as SKU and synthesized name."""
    cur = conn.cursor()
    inserted = 0

    for p in products:
        product_code = get_field(p, "Product_Code", "Product Code", "product_code")
        serial = get_field(p, "Product_Serial_Number", "Product Serial Number", "serial_number")
        category = get_field(p, "Category_Number", "Category Number", "category")

        if not product_code:
            continue

        sku = product_code  # Use Product_Code as SKU
        name = f"Product {sku}"  # Synthesize product name

        cur.execute("""
            INSERT INTO products (sku, name, category, unit_of_measure, weight_kg, barcode_value)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (sku) DO UPDATE SET
                name = EXCLUDED.name,
                category = EXCLUDED.category,
                barcode_value = EXCLUDED.barcode_value,
                updated_at = NOW()
        """, (sku, name, category if category else None, "EACH", None, serial if serial else None))
        inserted += 1

    conn.commit()
    cur.close()
    print(f"  ✓ Products seeded: {inserted} items (SKU = Product_Code)")
    return inserted


def seed_inventory(conn, products: list[dict], bin_ids: dict):
    """Create inventory records mapping products to bins with expected_qty = 1."""
    cur = conn.cursor()
    linked = 0

    for p in products:
        code = get_field(p, "Product_Code", "Product Code", "product_code")
        sku = code

        loc = extract_location_from_code(code)
        if not loc or not sku:
            continue

        a, rk, s, pos = loc["aisle"], loc["rack"], loc["shelf"], loc["product_pos"]
        bin_code = f"A{a}-RK{rk}-S{s}-B{pos}"
        bin_id = bin_ids.get(bin_code)

        if not bin_id:
            continue

        cur.execute("""
            INSERT INTO inventory (bin_id, sku, expected_qty)
            VALUES (%s, %s, %s)
            ON CONFLICT (bin_id, sku) DO UPDATE SET
                expected_qty = EXCLUDED.expected_qty,
                last_wms_sync = NOW(),
                updated_at = NOW()
        """, (bin_id, sku, 1))
        linked += 1

    conn.commit()
    cur.close()
    print(f"  ✓ Inventory records: {linked} product-bin mappings (expected_qty = 1)")


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  WAREOps — Warehouse Database Seeder")
    print("=" * 60)

    if not os.path.exists(XLSX_PATH):
        print(f"\n✗ ERROR: Spreadsheet not found: {XLSX_PATH}")
        print("  Make sure warehouse_database.xlsx is in the project root.")
        sys.exit(1)

    print(f"\n→ Reading spreadsheet: {XLSX_PATH}")
    products = parse_warehouse_data(XLSX_PATH)
    print(f"  Found {len(products)} products")

    print(f"\n→ Connecting to database...")
    print(f"  URL: {DATABASE_URL[:50]}...")

    try:
        conn = psycopg2.connect(DATABASE_URL)
    except Exception as e:
        print(f"\n✗ ERROR: Could not connect to database: {e}")
        print("  Set DATABASE_URL environment variable or check your PostgreSQL instance.")
        sys.exit(1)

    try:
        print(f"\n→ Seeding demo warehouse...")
        seed_warehouse(conn)

        print(f"\n→ Seeding topology (zones, aisles, racks, shelves, bins)...")
        bin_ids = seed_topology(conn, products)

        print(f"\n→ Seeding products...")
        seed_products(conn, products)

        print(f"\n→ Seeding inventory records...")
        seed_inventory(conn, products, bin_ids)

        print(f"\n{'=' * 60}")
        print(f"  ✓ Database seeding complete!")
        print(f"{'=' * 60}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()

