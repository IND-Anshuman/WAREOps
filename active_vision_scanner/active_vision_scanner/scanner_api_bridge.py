"""
scanner_api_bridge.py — Bridge module connecting the Active Vision Scanner to the
WAREOps backend API for remote operation.

This module provides:
  1. Product database fetching from the backend API
  2. Scan observation submission to the backend
  3. Mismatch alert creation via the backend
  4. Fallback to local xlsx if API is unreachable
"""
from __future__ import annotations

import os
import json
import logging
from datetime import datetime, timezone
from typing import Optional
from dataclasses import dataclass, field, asdict

logger = logging.getLogger("scanner_api_bridge")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    logger.warning("requests library not available — running in offline mode only")


@dataclass
class ScanObservation:
    """A single scan observation to submit to the backend."""
    warehouse_id: str
    robot_id: str
    bin_code: str
    decoded_qr: str
    observed_at: str
    detection_confidence: float = 0.0
    image_url: Optional[str] = None
    robot_coord_x: float = 0.0
    robot_coord_y: float = 0.0
    robot_coord_z: float = 0.0
    mission_id: Optional[str] = None


class ScannerAPIBridge:
    def __init__(
        self,
        api_url: Optional[str] = None,
        api_token: Optional[str] = None,
        warehouse_id: Optional[str] = None,
        robot_id: Optional[str] = None,
        timeout: float = 10.0,
    ):
        self.api_url = (api_url or os.environ.get("WAREOPS_API_URL", "")).rstrip("/")
        self.api_token = api_token or os.environ.get("WAREOPS_API_TOKEN", "")
        self.warehouse_id = warehouse_id or os.environ.get(
            "WAREOPS_WAREHOUSE_ID", "a1b2c3d4-e5f6-7890-abcd-ef1234567890"
        )
        self.robot_id = robot_id or os.environ.get(
            "WAREOPS_SCANNER_ROBOT_ID", "robot-1234-5678-90ab"
        )
        self.timeout = timeout
        self._online = bool(self.api_url and HAS_REQUESTS)
        self._session: Optional[requests.Session] = None
        self._bin_cache: dict[str, str] = {}

        if self._online:
            self._session = requests.Session()
            self._session.headers.update({
                "Content-Type": "application/json",
                "Accept": "application/json",
            })
            if self.api_token:
                self._session.headers["Authorization"] = f"Bearer {self.api_token}"

        logger.info(
            "ScannerAPIBridge initialized",
            extra={
                "api_url": self.api_url or "(offline)",
                "warehouse_id": self.warehouse_id,
                "robot_id": self.robot_id,
                "online": self._online,
            },
        )

    @property
    def is_online(self) -> bool:
        if not self._online or not self._session:
            return False
        try:
            resp = self._session.get(f"{self.api_url}/health", timeout=5)
            return resp.status_code == 200
        except Exception:
            return False

    def get_product_database(self) -> list[dict]:
        if self._online and self._session:
            try:
                resp = self._session.get(f"{self.api_url}/api/v1/products", timeout=self.timeout)
                resp.raise_for_status()
                products = resp.json()
                logger.info(f"Fetched {len(products)} products from backend API")
                return products
            except Exception as e:
                logger.warning(f"Failed to fetch products from API: {e}. Falling back to local.")
        return self._load_local_products()

    def _load_local_products(self) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not available for local fallback")
            return []

        xlsx_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "warehouse_database.xlsx",
        )
        if not os.path.exists(xlsx_path):
            logger.error(f"Local spreadsheet not found: {xlsx_path}")
            return []

        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        headers = []
        products = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            if not any(row):
                continue
            record = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    record[headers[col_idx]] = value
            
            products.append({
                "sku": record.get("Product_Code", ""),
                "name": f"Product {record.get('Product_Serial_Number', '')}",
                "category": record.get("Category_Number", ""),
                "weight_kg": 0.0,
                "barcode_value": record.get("QR_Code", ""),
            })

        wb.close()
        logger.info(f"Loaded {len(products)} products from local spreadsheet")
        return products

    def submit_observation(
        self,
        bin_code: str,
        decoded_qr: str,
        confidence: float = 0.0,
        image_url: Optional[str] = None,
        robot_x: float = 0.0,
        robot_y: float = 0.0,
        robot_z: float = 0.0,
        mission_id: Optional[str] = None,
    ) -> Optional[dict]:
        obs = ScanObservation(
            warehouse_id=self.warehouse_id,
            robot_id=self.robot_id,
            bin_code=bin_code,
            decoded_qr=decoded_qr,
            observed_at=datetime.now(timezone.utc).isoformat(),
            detection_confidence=confidence,
            image_url=image_url,
            robot_coord_x=robot_x,
            robot_coord_y=robot_y,
            robot_coord_z=robot_z,
            mission_id=mission_id,
        )

        if not self._online or not self._session:
            logger.info(f"[OFFLINE] Observation: {bin_code} → {decoded_qr} ({confidence:.0%})")
            return asdict(obs)

        try:
            # Observation endpoint expects ObservationBatch
            payload = {
                "observations": [asdict(obs)]
            }
            resp = self._session.post(
                f"{self.api_url}/api/v1/observations/batch",
                json=payload,
                timeout=self.timeout,
            )
            resp.raise_for_status()
            result = resp.json()
            logger.info(
                f"Observation submitted: {bin_code} → {decoded_qr} ({confidence:.0%})"
            )
            return result
        except Exception as e:
            logger.error(f"Failed to submit observation: {e}")
            return None

    def _resolve_bin_id(self, bin_code: str) -> Optional[str]:
        if bin_code in self._bin_cache:
            return self._bin_cache[bin_code]

        if not self._online or not self._session:
            return None

        try:
            resp = self._session.get(
                f"{self.api_url}/api/v1/warehouses/{self.warehouse_id}/topology",
                timeout=self.timeout,
            )
            resp.raise_for_status()
            topo = resp.json()
            # Topology: zones → aisles → racks → shelves → bins
            for zone in topo.get("zones", []):
                for aisle in zone.get("aisles", []):
                    for rack in aisle.get("racks", []):
                        for shelf in rack.get("shelves", []):
                            for bin_ in shelf.get("bins", []):
                                code = str(bin_.get("code", ""))
                                if code:
                                    self._bin_cache[code] = str(bin_["id"])
            return self._bin_cache.get(bin_code)
        except Exception as e:
            logger.error(f"Failed to resolve bin_id for {bin_code}: {e}")
            return None

    def submit_mismatch_alert(
        self,
        bin_code: str,
        expected_value: str,
        observed_value: Optional[str],
        mismatch_type: str = "MISPLACED",
        severity: str = "HIGH",
        confidence: float = 0.0,
    ) -> Optional[dict]:
        bin_id = self._resolve_bin_id(bin_code)
        
        if not self._online or not self._session or not bin_id:
            logger.warning(
                f"[OFFLINE] Alert: {mismatch_type} at {bin_code} "
                f"(expected={expected_value}, observed={observed_value})"
            )
            return None

        alert_data = {
            "warehouse_id": self.warehouse_id,
            "bin_id": bin_id,
            "alert_type": mismatch_type,
            "severity": severity,
            "title": f"{mismatch_type}: {bin_code}",
            "description": (
                f"Expected '{expected_value}' but observed "
                f"'{observed_value or 'EMPTY'}' at bin {bin_code}. "
                f"Confidence: {confidence:.0%}"
            ),
            "expected_value": expected_value,
            "observed_value": observed_value,
        }

        try:
            resp = self._session.post(
                f"{self.api_url}/api/v1/alerts",
                json=alert_data,
                timeout=self.timeout,
            )
            resp.raise_for_status()
            result = resp.json()
            logger.info(
                f"Alert created: {mismatch_type} at {bin_code}",
                extra={"alert_id": result.get("id")},
            )
            return result
        except Exception as e:
            logger.error(f"Failed to create alert: {e}")
            return None

    def send_robot_heartbeat(
        self,
        x: float,
        y: float,
        z: float = 0.0,
        yaw: float = 0.0,
        battery_pct: float = 100.0,
        status: str = "AUDITING",
        mission_id: Optional[str] = None
    ) -> bool:
        if not self._online or not self._session:
            return False

        # twin_router RobotHeartbeatRequest requires warehouse_id + robot_id in body
        heartbeat = {
            "warehouse_id": self.warehouse_id,
            "robot_id": self.robot_id,
            "x": x,
            "y": y,
            "z": z,
            "yaw": yaw,
            "battery_pct": battery_pct,
            "status": status,
            "mission_id": mission_id,
        }

        try:
            resp = self._session.post(
                f"{self.api_url}/api/v1/twin/robot-heartbeat",
                json=heartbeat,
                timeout=5,
            )
            return resp.status_code in (200, 201, 204)
        except Exception:
            return False

