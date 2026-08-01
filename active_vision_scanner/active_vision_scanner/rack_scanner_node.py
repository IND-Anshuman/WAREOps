#!/usr/bin/env python3
"""
Active Vision Rack Scanning System - ROS2 OpenCV Node
=====================================================

Network Setup:
    All devices connect to the SAME mobile hotspot:
    - Laptop (runs this ROS2 node + esp32_wifi_bridge)
    - ESP32-CAM (streams video over WiFi)
    - ESP32 Dev Kit V1 (servo controller, receives HTTP commands)

This node subscribes to a robot navigation topic to know when the bot has
reached a specific aisle/rack position. Upon receiving a "Reached" message,
it commands the ESP32 servo controller (via the esp32_wifi_bridge node) to
move the camera from its parked position (yaw=90, pitch=90) to the default
scanning position, automatically adjusts pitch so the full rack is in frame,
runs QR-code detection on each shelf region, then publishes "Succeeded" and
parks the camera again.

Physical Parameters (user-specified):
    Camera height from ground : 13.35 cm
    Rack height               : 49 cm
    Rack horizontal width     : 60 cm
    Camera distance from rack : 45 cm  (from design doc)
    Default orientation       : yaw=0°, pitch=90°

ROS2 Topics:
    /bot_status        (std_msgs/String)  - Subscribe & Publish
        Robot publishes  : "Reached Aisle X, RackY"
        This node publishes: "Succeeded Aisle X, RackY"
    /camera_servo_cmd  (std_msgs/String)  - Publish servo commands as JSON
        → esp32_wifi_bridge forwards to ESP32 via HTTP POST
    /camera_system     (std_msgs/String)  - Subscribe; "stop" → shutdown
"""

import rclpy
from rclpy.node import Node
from std_msgs.msg import String

import cv2
import numpy as np
import json
import time
import threading
import os
import csv
import re
import openpyxl
from enum import Enum, auto

from active_vision_scanner.scanner_api_bridge import ScannerAPIBridge


# ---------------------------------------------------------------------------
# Constants – Physical Setup
# ---------------------------------------------------------------------------
CAMERA_HEIGHT_CM = 13.35          # Camera height from ground (cm)
RACK_HEIGHT_CM = 49.0             # Total rack height (cm)
RACK_HORIZONTAL_CM = 60.0         # Rack horizontal width (cm)
CAMERA_DISTANCE_CM = 40.0         # Distance from camera to rack (cm, approximately 40cm)
NUM_SHELVES = 4                   # Number of shelves on the rack

# Camera FOV (OV3660 approximate)
HFOV_DEG = 66.0                   # Horizontal FOV (degrees)
VFOV_DEG = 52.0                   # Vertical FOV (degrees)

# Frame resolution
FRAME_WIDTH = 640
FRAME_HEIGHT = 480

# Shelf pixel regions (4 equal horizontal bands)
SHELF_REGIONS = {
    1: (0, FRAME_HEIGHT // NUM_SHELVES),
    2: (FRAME_HEIGHT // NUM_SHELVES, 2 * FRAME_HEIGHT // NUM_SHELVES),
    3: (2 * FRAME_HEIGHT // NUM_SHELVES, 3 * FRAME_HEIGHT // NUM_SHELVES),
    4: (3 * FRAME_HEIGHT // NUM_SHELVES, FRAME_HEIGHT),
}

# Servo angles
PARKED_YAW = 90       # Parked yaw (looking sideways / stowed)
PARKED_PITCH = 90     # Parked pitch (looking straight ahead / stowed)
DEFAULT_YAW = 0       # Default scanning yaw
DEFAULT_PITCH = 90    # Default scanning pitch (will be auto-adjusted)

# Pitch adjustment
PITCH_STEP_DEG = 2    # Degrees per adjustment step
PITCH_MIN = 50        # Minimum allowed pitch
PITCH_MAX = 140       # Maximum allowed pitch

# ESP32-CAM stream URL — all devices on same mobile hotspot
# Update this IP to match your ESP32-CAM's IP on the hotspot network
ESP32_CAM_STREAM_URL = "http://192.168.43.100:81/stream"

# Servo movement speed (degrees per command for ESP32 trajectory)
SERVO_SPEED = 20


# ---------------------------------------------------------------------------
# State machine
# ---------------------------------------------------------------------------
class ScannerState(Enum):
    IDLE = auto()              # Waiting for a "Reached" message
    MOVING_TO_SCAN = auto()    # Moving servos from parked → default scan pos
    ADJUSTING_PITCH = auto()   # Auto-adjusting pitch to frame the rack
    SCANNING = auto()          # QR detection in progress
    MOVING_TO_PARK = auto()    # Moving servos back to parked position
    SHUTDOWN = auto()          # Shutting down


class RackScannerNode(Node):
    """ROS2 node that orchestrates rack scanning with OpenCV + servo control."""

    def __init__(self):
        super().__init__('rack_scanner_node')

        # ------- Parameters (overridable via ROS2 params) -------
        self.declare_parameter('esp32_cam_url', ESP32_CAM_STREAM_URL)
        self.declare_parameter('camera_height_cm', CAMERA_HEIGHT_CM)
        self.declare_parameter('rack_height_cm', RACK_HEIGHT_CM)
        self.declare_parameter('rack_horizontal_cm', RACK_HORIZONTAL_CM)
        self.declare_parameter('camera_distance_cm', CAMERA_DISTANCE_CM)
        self.declare_parameter('num_shelves', NUM_SHELVES)
        self.declare_parameter('database_path', os.path.join(os.getcwd(), 'warehouse_database.xlsx'))
        self.declare_parameter('scanned_inventory_path', os.path.join(os.getcwd(), 'scanned_inventory.csv'))
        self.declare_parameter('mismatch_log_path', os.path.join(os.getcwd(), 'mismatch_log.csv'))

        self.cam_url = self.get_parameter('esp32_cam_url').value
        self.camera_height = self.get_parameter('camera_height_cm').value
        self.rack_height = self.get_parameter('rack_height_cm').value
        self.rack_horizontal = self.get_parameter('rack_horizontal_cm').value
        self.camera_distance = self.get_parameter('camera_distance_cm').value
        self.num_shelves = self.get_parameter('num_shelves').value
        self.db_path = self.get_parameter('database_path').value
        self.scanned_inv_path = self.get_parameter('scanned_inventory_path').value
        self.mismatch_log_path = self.get_parameter('mismatch_log_path').value

        # ------- Publishers -------
        self.bot_status_pub = self.create_publisher(String, '/bot_status', 10)
        self.servo_cmd_pub = self.create_publisher(String, '/camera_servo_cmd', 10)

        # ------- Subscribers -------
        self.bot_status_sub = self.create_subscription(
            String, '/bot_status', self.bot_status_callback, 10
        )
        self.camera_system_sub = self.create_subscription(
            String, '/camera_system', self.camera_system_callback, 10
        )

        # ------- Internal state -------
        self.state = ScannerState.IDLE
        self.current_yaw = PARKED_YAW
        self.current_pitch = PARKED_PITCH
        self.current_location = ""        # e.g. "Aisle_1/Row_1/Rack_1"
        self.scanned_aisle = ""
        self.scanned_row = ""
        self.scanned_rack = ""
        self.scan_results = []            # Detected products
        self.cap = None                   # OpenCV VideoCapture
        self.lock = threading.Lock()
        self.shutdown_event = threading.Event()

        # QR detector
        self.qr_detector = cv2.QRCodeDetector()

        # Scanner API Bridge for WAREOps backend
        self.api_bridge = ScannerAPIBridge()

        # Load QR database
        self.database = {}
        self.load_database()

        # Timer for the main processing loop (10 Hz)
        self.create_timer(0.1, self.process_loop)

        self.get_logger().info(
            f"Rack Scanner Node started.\n"
            f"  Camera height : {self.camera_height} cm\n"
            f"  Rack height   : {self.rack_height} cm\n"
            f"  Rack width    : {self.rack_horizontal} cm\n"
            f"  Shelves       : {self.num_shelves}\n"
            f"  Parked pos    : yaw={PARKED_YAW}°, pitch={PARKED_PITCH}°\n"
            f"  Default scan  : yaw={DEFAULT_YAW}°, pitch={DEFAULT_PITCH}°\n"
            f"  Database file : {self.db_path}\n"
            f"  ESP32 stream  : {self.cam_url}"
        )

    def load_database(self):
        self.get_logger().info(f"Loading QR database...")
        
        # Try fetching from API first
        api_products = self.api_bridge.get_product_database()
        if api_products:
            self.database.clear()
            for p in api_products:
                # Backend returns: sku, name, category, weight_kg, barcode_value
                # We map it to the expected structure
                qr = p.get('barcode_value', p.get('sku'))
                if qr:
                    self.database[qr] = {
                        'product_code': p.get('sku'),
                        'serial_number': p.get('barcode_value'),
                        'category_number': p.get('category'),
                        # API doesn't provide specific shelf assignments, so we assume
                        # matching SKU means correct for now, or fallback to local
                        'aisle': 'A1',
                        'row': 'R1',
                        'rack': 'RK1',
                        'shelf': 'S1'
                    }
            self.get_logger().info(f"Successfully loaded {len(self.database)} entries from API database.")
            return

        # Fallback to local Excel
        if not os.path.exists(self.db_path):
            self.get_logger().error(f"Excel database file not found at {self.db_path}!")
            return
        
        try:
            wb = openpyxl.load_workbook(self.db_path)
            sheet = wb.active
            count = 0
            for r in range(2, sheet.max_row + 1):
                qr = sheet.cell(r, 1).value
                if qr:
                    qr = str(qr).strip()
                    self.database[qr] = {
                        'product_code': sheet.cell(r, 2).value,
                        'serial_number': sheet.cell(r, 3).value,
                        'category_number': sheet.cell(r, 4).value,
                        'aisle': sheet.cell(r, 5).value,
                        'row': sheet.cell(r, 6).value,
                        'rack': sheet.cell(r, 7).value,
                        'shelf': sheet.cell(r, 8).value
                    }
                    count += 1
            self.get_logger().info(f"Successfully loaded {count} entries from database.")
        except Exception as e:
            self.get_logger().error(f"Error loading Excel database: {e}")

    # ===========================================================================
    # ROS2 Callbacks
    # ===========================================================================
    def bot_status_callback(self, msg: String):
        """Handle messages on /bot_status."""
        data = msg.data.strip()
        self.get_logger().info(f"[/bot_status] Received: '{data}'")

        # Only react to "Reached ..." messages while idle
        if data.lower().startswith("reached") and self.state == ScannerState.IDLE:
            loc_str = data.replace("Reached", "").strip()
            parsed_aisle, parsed_row, parsed_rack = self.parse_location(loc_str)
            if parsed_aisle and parsed_row and parsed_rack:
                self.current_location = loc_str
                self.scanned_aisle = parsed_aisle
                self.scanned_row = parsed_row
                self.scanned_rack = parsed_rack
                self.get_logger().info(
                    f"Bot reached '{self.current_location}' -> Aisle: {self.scanned_aisle}, Row: {self.scanned_row}, Rack: {self.scanned_rack}"
                )
                self.state = ScannerState.MOVING_TO_SCAN
                
                # Send heartbeat to twin
                self.api_bridge.send_robot_heartbeat(
                    x=1.0, y=1.0, z=0.0, yaw=self.current_yaw, status="MOVING_TO_SCAN"
                )
            else:
                self.get_logger().error(f"Failed to parse Aisle/Row/Rack from location: '{loc_str}'")

    def parse_location(self, loc_str):
        # Match e.g., "Aisle_1/Row_1/Rack_1" or "Aisle1/Row1/Rack1"
        match = re.search(r'Aisle_?(\d+)/Row_?(\d+)/Rack_?(\d+)', loc_str, re.IGNORECASE)
        if match:
            return f"A{match.group(1)}", f"R{match.group(2)}", f"RK{match.group(3)}"
        return None, None, None

    def camera_system_callback(self, msg: String):
        """Handle messages on /camera_system — 'stop' triggers shutdown."""
        data = msg.data.strip().lower()
        self.get_logger().info(f"[/camera_system] Received: '{data}'")

        if data == "stop":
            self.get_logger().warn("Received STOP command — shutting down.")
            self.state = ScannerState.SHUTDOWN
            self.shutdown_event.set()

    # ===========================================================================
    # Servo Command Helpers
    # ===========================================================================
    def send_servo_command(self, pitch: float, yaw: float, speed: int = SERVO_SPEED):
        """Publish a JSON servo command on /camera_servo_cmd."""
        cmd = {
            "cmd": "MOVE",
            "pitch": int(pitch),
            "yaw": int(yaw),
            "speed": speed,
        }
        msg = String()
        msg.data = json.dumps(cmd)
        self.servo_cmd_pub.publish(msg)
        self.current_pitch = pitch
        self.current_yaw = yaw
        self.get_logger().info(f"Servo CMD → pitch={pitch}°, yaw={yaw}°, speed={speed}")

    def send_home_command(self):
        """Send HOME command to ESP32."""
        cmd = {"cmd": "HOME"}
        msg = String()
        msg.data = json.dumps(cmd)
        self.servo_cmd_pub.publish(msg)
        self.get_logger().info("Servo CMD → HOME")

    def send_stop_command(self):
        """Send STOP command to ESP32."""
        cmd = {"cmd": "STOP"}
        msg = String()
        msg.data = json.dumps(cmd)
        self.servo_cmd_pub.publish(msg)
        self.get_logger().info("Servo CMD → STOP")

    # ===========================================================================
    # Camera Helpers
    # ===========================================================================
    def open_camera(self) -> bool:
        """Open the ESP32-CAM video stream."""
        if self.cap is not None and self.cap.isOpened():
            return True
        self.get_logger().info(f"Opening camera stream: {self.cam_url}")
        self.cap = cv2.VideoCapture(self.cam_url)
        if not self.cap.isOpened():
            self.get_logger().error("Failed to open camera stream!")
            return False
        self.get_logger().info("Camera stream opened successfully.")
        return True

    def close_camera(self):
        """Release the camera stream."""
        if self.cap is not None:
            self.cap.release()
            self.cap = None
            self.get_logger().info("Camera stream closed.")

    def grab_frame(self):
        """Grab a single frame from the camera. Returns (ok, frame)."""
        if self.cap is None or not self.cap.isOpened():
            return False, None
        ret, frame = self.cap.read()
        if not ret:
            self.get_logger().warn("Failed to grab frame.")
            return False, None
        # Resize to expected resolution
        frame = cv2.resize(frame, (FRAME_WIDTH, FRAME_HEIGHT))
        return True, frame

    # ===========================================================================
    # Rack Framing / Pitch Adjustment
    # ===========================================================================
    def is_rack_fully_visible(self, frame) -> tuple:
        """
        Determine whether the full rack is visible in the frame.

        Strategy: Convert to grayscale, apply edge detection, and check if
        significant edges exist in the top and bottom regions of the frame.
        The top ~10% and bottom ~10% strips are checked.

        Returns (top_visible: bool, bottom_visible: bool).
        """
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150)

        top_strip = edges[0:FRAME_HEIGHT // 10, :]
        bottom_strip = edges[9 * FRAME_HEIGHT // 10:, :]

        top_edge_density = np.count_nonzero(top_strip) / top_strip.size
        bottom_edge_density = np.count_nonzero(bottom_strip) / bottom_strip.size

        # Thresholds — if there are very few edges, the shelf edge is likely
        # outside the frame.
        EDGE_THRESHOLD = 0.02
        top_visible = top_edge_density > EDGE_THRESHOLD
        bottom_visible = bottom_edge_density > EDGE_THRESHOLD

        return top_visible, bottom_visible

    def adjust_pitch_for_rack(self, max_steps: int = 20) -> bool:
        """
        Automatically adjust pitch so the entire rack is in frame.

        Returns True if adjustment succeeded, False on failure/timeout.
        """
        self.get_logger().info("Starting automatic pitch adjustment...")
        for step in range(max_steps):
            ok, frame = self.grab_frame()
            if not ok:
                self.get_logger().warn("Cannot grab frame during pitch adjustment.")
                time.sleep(0.3)
                continue

            top_vis, bottom_vis = self.is_rack_fully_visible(frame)
            self.get_logger().info(
                f"  Step {step}: top_visible={top_vis}, bottom_visible={bottom_vis}, "
                f"pitch={self.current_pitch}°"
            )

            if top_vis and bottom_vis:
                self.get_logger().info(
                    f"Rack fully in frame at pitch={self.current_pitch}°."
                )
                return True

            if not top_vis and bottom_vis:
                # Top shelf is out of frame → decrease pitch (tilt up)
                new_pitch = max(PITCH_MIN, self.current_pitch - PITCH_STEP_DEG)
                self.send_servo_command(new_pitch, self.current_yaw)
            elif top_vis and not bottom_vis:
                # Bottom shelf is out of frame → increase pitch (tilt down)
                new_pitch = min(PITCH_MAX, self.current_pitch + PITCH_STEP_DEG)
                self.send_servo_command(new_pitch, self.current_yaw)
            else:
                # Neither visible — try decreasing pitch (camera too close or wrong angle)
                new_pitch = max(PITCH_MIN, self.current_pitch - PITCH_STEP_DEG)
                self.send_servo_command(new_pitch, self.current_yaw)

            # Wait for servo to settle
            time.sleep(0.5)

        self.get_logger().warn("Pitch adjustment did not converge within max steps.")
        return False

    # ===========================================================================
    # QR Code Detection
    # ===========================================================================
    def determine_shelf(self, center_y: float, scan_mode: str = 'FULL') -> int:
        """
        Determine which shelf a QR code belongs to based on its center_y and current scan mode.
        - FULL (Centered view):
            0..120 -> S1, 120..240 -> S2, 240..360 -> S3, 360..480 -> S4
        - TOP_SWEEP (Tilted UP to view top half of rack):
            Upper half of screen -> S1, Lower half -> S2
        - BOTTOM_SWEEP (Tilted DOWN to view bottom half of rack):
            Upper half of screen -> S3, Lower half -> S4
        """
        if scan_mode == 'TOP_SWEEP':
            return 1 if center_y < (FRAME_HEIGHT / 2.0) else 2
        elif scan_mode == 'BOTTOM_SWEEP':
            return 3 if center_y < (FRAME_HEIGHT / 2.0) else 4
        else: # FULL mode
            for shelf_id, (y_min, y_max) in SHELF_REGIONS.items():
                if y_min <= center_y < y_max:
                    return shelf_id
        return -1

    def detect_qr_codes(self, frame, scan_mode: str = 'FULL') -> list:
        """
        Detect and decode all QR codes in the frame.

        Returns a list of dicts with keys:
            data, shelf_detected, center_x, center_y, bbox
        """
        detections = []

        # Use OpenCV's QRCodeDetector for multi-code detection
        try:
            retval, decoded_info, points, _ = self.qr_detector.detectAndDecodeMulti(frame)
            if retval and decoded_info is not None:
                for i, data in enumerate(decoded_info):
                    if data:  # Non-empty decoded string
                        bbox = points[i]
                        # Compute center
                        cx = np.mean(bbox[:, 0])
                        cy = np.mean(bbox[:, 1])
                        shelf = self.determine_shelf(cy, scan_mode=scan_mode)

                        # Try to parse JSON from QR data
                        product_info = self._parse_qr_data(data)

                        detections.append({
                            'raw_data': data,
                            'product_info': product_info,
                            'shelf_detected': shelf,
                            'center_x': float(cx),
                            'center_y': float(cy),
                            'bbox': bbox.tolist(),
                        })
                        self.get_logger().info(
                            f"QR detected [{scan_mode}]: shelf={shelf}, center=({cx:.0f},{cy:.0f}), "
                            f"data='{data[:60]}...'"
                        )
        except Exception as e:
            # Fallback: single QR detection
            data, bbox, _ = self.qr_detector.detectAndDecode(frame)
            if data and bbox is not None:
                cx = np.mean(bbox[0][:, 0])
                cy = np.mean(bbox[0][:, 1])
                shelf = self.determine_shelf(cy, scan_mode=scan_mode)
                product_info = self._parse_qr_data(data)
                detections.append({
                    'raw_data': data,
                    'product_info': product_info,
                    'shelf_detected': shelf,
                    'center_x': float(cx),
                    'center_y': float(cy),
                    'bbox': bbox[0].tolist(),
                })
                self.get_logger().info(
                    f"QR detected (single) [{scan_mode}]: shelf={shelf}, data='{data[:60]}'"
                )

        return detections

    def _parse_qr_data(self, data: str) -> dict:
        """Try to parse QR data as JSON; fall back to raw string."""
        try:
            return json.loads(data)
        except (json.JSONDecodeError, TypeError):
            return {'raw': data}

    def run_scan(self) -> list:
        """
        Run a full scan: grab multiple frames at centered pitch angle, and optionally
        perform micro-pitch tilt sweeps (tilt UP for S1/S2, tilt DOWN for S3/S4) to ensure
        complete coverage at close distances (~40cm).
        """
        all_detections = {}

        # Phase 1: Centered Scan (10 frames)
        self.get_logger().info("Phase 1: Centered scan (FULL frame)...")
        for i in range(10):
            ok, frame = self.grab_frame()
            if not ok:
                time.sleep(0.1)
                continue

            detections = self.detect_qr_codes(frame, scan_mode='FULL')
            for det in detections:
                key = det['raw_data']
                if key not in all_detections:
                    all_detections[key] = det
                    all_detections[key]['count'] = 1
                else:
                    all_detections[key]['count'] += 1
            time.sleep(0.08)

        # Check which shelves were detected
        detected_shelves = set(d['shelf_detected'] for d in all_detections.values())
        
        # Phase 2: Active Pitch Tilt Sweep if any shelf missing (e.g. at 40cm distance)
        if len(detected_shelves) < 4:
            self.get_logger().info(f"Shelves detected so far: {detected_shelves}. Performing active pitch sweep...")
            center_pitch = self.current_pitch

            # Tilt UP by 12 deg to target Top Shelves (S1 & S2)
            top_pitch = max(PITCH_MIN, center_pitch - 12)
            self.send_servo_command(top_pitch, self.current_yaw)
            time.sleep(0.4)
            for _ in range(5):
                ok, frame = self.grab_frame()
                if ok:
                    for det in self.detect_qr_codes(frame, scan_mode='TOP_SWEEP'):
                        key = det['raw_data']
                        if key not in all_detections:
                            all_detections[key] = det
                            all_detections[key]['count'] = 1
                time.sleep(0.08)

            # Tilt DOWN by 12 deg to target Bottom Shelves (S3 & S4)
            bot_pitch = min(PITCH_MAX, center_pitch + 12)
            self.send_servo_command(bot_pitch, self.current_yaw)
            time.sleep(0.4)
            for _ in range(5):
                ok, frame = self.grab_frame()
                if ok:
                    for det in self.detect_qr_codes(frame, scan_mode='BOTTOM_SWEEP'):
                        key = det['raw_data']
                        if key not in all_detections:
                            all_detections[key] = det
                            all_detections[key]['count'] = 1
                time.sleep(0.08)

            # Return camera to centered pitch angle
            self.send_servo_command(center_pitch, self.current_yaw)
            time.sleep(0.3)

        results = list(all_detections.values())
        self.get_logger().info(
            f"Scan complete — {len(results)} unique QR codes detected."
        )
        for r in results:
            self.get_logger().info(
                f"  Product: {r.get('product_info', {})}, "
                f"Shelf: {r['shelf_detected']}, "
                f"Seen: {r['count']}x"
            )
        return results

    # ===========================================================================
    # Visualization (optional, runs in a thread)
    # ===========================================================================
    def draw_overlay(self, frame, detections):
        """Draw shelf regions and QR detections on the frame."""
        overlay = frame.copy()

        # Draw shelf dividers
        for shelf_id, (y_min, y_max) in SHELF_REGIONS.items():
            cv2.line(overlay, (0, y_min), (FRAME_WIDTH, y_min), (0, 255, 0), 1)
            cv2.putText(
                overlay, f"Shelf {shelf_id}",
                (10, y_min + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2
            )

        # Draw QR bounding boxes
        for det in detections:
            pts = np.array(det['bbox'], dtype=np.int32)
            cv2.polylines(overlay, [pts], True, (0, 0, 255), 2)
            cx, cy = int(det['center_x']), int(det['center_y'])
            cv2.circle(overlay, (cx, cy), 5, (255, 0, 0), -1)
            label = f"S{det['shelf_detected']}"
            cv2.putText(
                overlay, label, (cx + 10, cy),
                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2
            )

        # Status bar
        cv2.putText(
            overlay,
            f"State: {self.state.name} | Pitch: {self.current_pitch} | Yaw: {self.current_yaw}",
            (10, FRAME_HEIGHT - 15), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1
        )

        return overlay

    # ===========================================================================
    # Main Processing Loop (called by ROS2 timer)
    # ===========================================================================
    def process_loop(self):
        """State machine — runs at 10 Hz."""

        # ---- SHUTDOWN ----
        if self.state == ScannerState.SHUTDOWN:
            self.get_logger().info("Shutdown state — cleaning up...")
            self.send_stop_command()
            self.close_camera()
            cv2.destroyAllWindows()
            rclpy.shutdown()
            return

        # ---- IDLE ----
        if self.state == ScannerState.IDLE:
            # Nothing to do; waiting for /bot_status callback
            return

        # ---- MOVING TO SCAN POSITION ----
        if self.state == ScannerState.MOVING_TO_SCAN:
            self.get_logger().info("Moving camera from parked to default scan position...")

            # Open camera
            if not self.open_camera():
                self.get_logger().error("Cannot open camera — aborting scan.")
                self.state = ScannerState.IDLE
                return

            # Command servos to default scan orientation
            self.send_servo_command(DEFAULT_PITCH, DEFAULT_YAW)

            # Wait for servos to reach position
            # (The ESP32 does trajectory planning, so we wait proportionally)
            travel_yaw = abs(PARKED_YAW - DEFAULT_YAW)
            travel_pitch = abs(PARKED_PITCH - DEFAULT_PITCH)
            max_travel = max(travel_yaw, travel_pitch)
            wait_time = max_travel * 0.05 + 0.5  # ~50ms per degree + buffer
            self.get_logger().info(f"Waiting {wait_time:.1f}s for servos to settle...")
            time.sleep(wait_time)

            self.state = ScannerState.ADJUSTING_PITCH

        # ---- ADJUSTING PITCH ----
        elif self.state == ScannerState.ADJUSTING_PITCH:
            success = self.adjust_pitch_for_rack()
            if not success:
                self.get_logger().warn(
                    "Pitch adjustment failed — proceeding with current angle."
                )
            self.state = ScannerState.SCANNING

        # ---- SCANNING ----
        elif self.state == ScannerState.SCANNING:
            self.scan_results = self.run_scan()

            # Match scanned QRs against the database and log to CSV
            self.process_scan_results(self.scan_results)

            # Publish success
            result_msg = String()
            result_msg.data = f"Succeeded {self.current_location}"
            self.bot_status_pub.publish(result_msg)
            self.get_logger().info(f"Published: '{result_msg.data}'")

            self.state = ScannerState.MOVING_TO_PARK

        # ---- MOVING TO PARK ----
        elif self.state == ScannerState.MOVING_TO_PARK:
            self.get_logger().info("Moving camera back to parked position...")
            self.send_servo_command(PARKED_PITCH, PARKED_YAW)

            # Wait for servos
            travel_yaw = abs(self.current_yaw - PARKED_YAW)
            travel_pitch = abs(self.current_pitch - PARKED_PITCH)
            max_travel = max(travel_yaw, travel_pitch)
            wait_time = max_travel * 0.05 + 0.5
            time.sleep(wait_time)

            self.close_camera()
            self.get_logger().info("Scan cycle complete — returning to IDLE.")
            self.state = ScannerState.IDLE

    def process_scan_results(self, results):
        self.get_logger().info("Processing scan results against database...")
        
        # Prepare list for this position's correct scans
        position_correct_scans = []

        for r in results:
            qr_data = r['raw_data'].strip()
            shelf_num = r['shelf_detected']
            shelf_code = f"S{shelf_num}" if shelf_num in [1, 2, 3, 4] else "UNKNOWN"
            
            if qr_data in self.database:
                db_entry = self.database[qr_data]
                expected_aisle = db_entry['aisle']
                expected_row = db_entry['row']
                expected_rack = db_entry['rack']
                expected_shelf = db_entry['shelf']
                
                is_correct = (
                    expected_aisle == self.scanned_aisle and
                    expected_row == self.scanned_row and
                    expected_rack == self.scanned_rack and
                    expected_shelf == shelf_code
                )
                
                if is_correct:
                    self.get_logger().info(f"Match: QR code '{qr_data}' is correctly placed at {self.current_location}/{shelf_code}")
                    self.log_correct_scan(qr_data, db_entry, shelf_code)
                    position_correct_scans.append({
                        'QR_Code': qr_data,
                        'Product_Code': db_entry['product_code'],
                        'Product_Serial_Number': db_entry['serial_number'],
                        'Category_Number': db_entry['category_number'],
                        'Shelf': shelf_code
                    })
                else:
                    self.get_logger().warn(
                        f"Mismatch! At position {self.scanned_aisle}/{self.scanned_row}/{self.scanned_rack}/{shelf_code}, "
                        f"QR: '{qr_data}' ({db_entry['product_code']}) is present but it should be in "
                        f"Aisle: {expected_aisle}, Row: {expected_row}, Rack: {expected_rack}, Shelf: {expected_shelf}"
                    )
                    self.log_mismatch_scan(qr_data, db_entry, shelf_code, expected_aisle, expected_row, expected_rack, expected_shelf)
            else:
                self.get_logger().error(f"Unknown QR code scanned: '{qr_data}' at shelf {shelf_code}!")
                self.log_unknown_scan(qr_data, shelf_code)

        # Compile that position's all data in a position-specific CSV file
        if position_correct_scans:
            self.compile_position_csv(position_correct_scans)

    def log_correct_scan(self, qr, db_entry, shelf):
        file_exists = os.path.exists(self.scanned_inv_path)
        try:
            with open(self.scanned_inv_path, 'a', newline='') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow([
                        'Timestamp', 'Scanned_Location', 'Aisle', 'Row', 'Rack', 'Shelf',
                        'QR_Code', 'Product_Code', 'Product_Serial_Number', 'Category_Number'
                    ])
                writer.writerow([
                    time.strftime('%Y-%m-%d %H:%M:%S'),
                    self.current_location,
                    self.scanned_aisle,
                    self.scanned_row,
                    self.scanned_rack,
                    shelf,
                    qr,
                    db_entry['product_code'],
                    db_entry['serial_number'],
                    db_entry['category_number']
                ])
                
            # Submit to backend API
            bin_code = f"{self.scanned_aisle}-{self.scanned_row}-{self.scanned_rack}-{shelf}"
            self.api_bridge.submit_observation(
                bin_code=bin_code,
                decoded_qr=qr,
                confidence=0.98,
            )
        except Exception as e:
            self.get_logger().error(f"Failed to log correct scan to CSV: {e}")

    def log_mismatch_scan(self, qr, db_entry, shelf, exp_a, exp_row, exp_rk, exp_s):
        file_exists = os.path.exists(self.mismatch_log_path)
        message = (
            f"At position Aisle_{self.scanned_aisle[1:]}/Row_{self.scanned_row[1:]}/Rack_{self.scanned_rack[2:]}/{shelf} "
            f"this QR product ({db_entry['product_code']}) is there but it should be in position "
            f"Aisle_{exp_a[1:]}/Row_{exp_row[1:]}/Rack_{exp_rk[2:]}/{exp_s}"
        )
        try:
            with open(self.mismatch_log_path, 'a', newline='') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow([
                        'Timestamp', 'Current_Location', 'Current_Aisle', 'Current_Row', 'Current_Rack', 'Current_Shelf',
                        'QR_Code', 'Product_Code', 'Product_Serial_Number', 'Category_Number',
                        'Expected_Aisle', 'Expected_Row', 'Expected_Rack', 'Expected_Shelf', 'Message'
                    ])
                writer.writerow([
                    time.strftime('%Y-%m-%d %H:%M:%S'),
                    self.current_location,
                    self.scanned_aisle,
                    self.scanned_row,
                    self.scanned_rack,
                    shelf,
                    qr,
                    db_entry['product_code'],
                    db_entry['serial_number'],
                    db_entry['category_number'],
                    exp_a,
                    exp_row,
                    exp_rk,
                    exp_s,
                    message
                ])
                
            # Submit alert to backend API
            bin_code = f"{self.scanned_aisle}-{self.scanned_row}-{self.scanned_rack}-{shelf}"
            self.api_bridge.submit_mismatch_alert(
                bin_code=bin_code,
                expected_value=f"Location: {exp_a}/{exp_row}/{exp_rk}/{exp_s}",
                observed_value=db_entry['product_code'],
                mismatch_type="MISPLACED",
                severity="HIGH",
                confidence=0.95,
            )
        except Exception as e:
            self.get_logger().error(f"Failed to log mismatch scan to CSV: {e}")

    def log_unknown_scan(self, qr, shelf):
        file_exists = os.path.exists(self.mismatch_log_path)
        message = (
            f"At position Aisle_{self.scanned_aisle[1:]}/Row_{self.scanned_row[1:]}/Rack_{self.scanned_rack[2:]}/{shelf} "
            f"unknown QR product ({qr}) is present."
        )
        try:
            with open(self.mismatch_log_path, 'a', newline='') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow([
                        'Timestamp', 'Current_Location', 'Current_Aisle', 'Current_Row', 'Current_Rack', 'Current_Shelf',
                        'QR_Code', 'Product_Code', 'Product_Serial_Number', 'Category_Number',
                        'Expected_Aisle', 'Expected_Row', 'Expected_Rack', 'Expected_Shelf', 'Message'
                    ])
                writer.writerow([
                    time.strftime('%Y-%m-%d %H:%M:%S'),
                    self.current_location,
                    self.scanned_aisle,
                    self.scanned_row,
                    self.scanned_rack,
                    shelf,
                    qr,
                    'UNKNOWN',
                    'UNKNOWN',
                    'UNKNOWN',
                    'UNKNOWN',
                    'UNKNOWN',
                    'UNKNOWN',
                    'UNKNOWN',
                    message
                ])
                
            # Submit alert to backend API
            bin_code = f"{self.scanned_aisle}-{self.scanned_row}-{self.scanned_rack}-{shelf}"
            self.api_bridge.submit_mismatch_alert(
                bin_code=bin_code,
                expected_value="Known Product",
                observed_value=qr,
                mismatch_type="UNKNOWN_QR",
                severity="MEDIUM",
                confidence=0.9,
            )
        except Exception as e:
            self.get_logger().error(f"Failed to log unknown scan to CSV: {e}")

    def compile_position_csv(self, correct_scans):
        safe_loc_name = self.current_location.replace('/', '_').replace(' ', '_')
        pos_filename = os.path.join(os.getcwd(), f"scanned_{safe_loc_name}.csv")
        
        self.get_logger().info(f"Compiling correct position scans to {pos_filename}...")
        try:
            with open(pos_filename, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['QR_Code', 'Product_Code', 'Product_Serial_Number', 'Category_Number', 'Shelf'])
                for scan in correct_scans:
                    writer.writerow([
                        scan['QR_Code'],
                        scan['Product_Code'],
                        scan['Product_Serial_Number'],
                        scan['Category_Number'],
                        scan['Shelf']
                    ])
        except Exception as e:
            self.get_logger().error(f"Failed to write position specific CSV: {e}")


# ===========================================================================
# Entry point
# ===========================================================================
def main(args=None):
    rclpy.init(args=args)
    node = RackScannerNode()

    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        node.get_logger().info("Keyboard interrupt — shutting down.")
    finally:
        node.close_camera()
        cv2.destroyAllWindows()
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()


if __name__ == '__main__':
    main()
